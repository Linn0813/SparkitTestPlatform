from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Optional

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.case import CaseModule, CasePriority, TestCase
from app.models.project import ProjectMember
from app.models.requirement import Requirement
from app.models.template import ProjectFieldTemplate, TemplateScene
from app.models.user import User
from app.services.case_module_paths import load_project_modules
from app.services.case_module_resolve import ImportModuleResolver
from app.services.field_validator import load_project_member_user_ids, validate_custom_fields
from app.services.links import set_case_requirements, validate_requirement_ids
from app.services.template_fields import RESERVED_FIELD_NAMES_BY_SCENE

# 与前端 constants/systemFields.ts 中 SYSTEM_CASE_FIELDS 顺序一致
CASE_SYSTEM_IMPORT_COLUMNS: list[tuple[str, Optional[str]]] = [
    ("模块", "module_path"),
    ("用例标题", "title"),
    ("优先级", "priority"),
    ("前置条件", "precondition"),
    ("步骤", "step_text"),
    ("预期结果", "expected_result"),
    ("关联需求", "requirement_ids"),
]

CASE_RESERVED_FIELD_NAMES = RESERVED_FIELD_NAMES_BY_SCENE["functional_case"]

# 表头别名（兼容 MeterSphere / 手工表格常用列名）
HEADER_ALIASES: dict[str, str] = {
    "标题": "用例标题",
    "用例名称": "用例标题",
    "名称": "用例标题",
    "所属模块": "模块",
    "模块路径": "模块",
    "步骤描述": "步骤",
    "步骤说明": "步骤",
    "预期": "预期结果",
    "需求": "关联需求",
    "关联需求编号": "关联需求",
}

SKIP_FIELD_TYPES = frozenset({"attachment", "requirement_link", "plan_link", "version_link", "status", "module"})

TAG_SPLIT = re.compile(r"[,，;；\n]+")
MULTI_SPLIT = re.compile(r"[,，;；\n]+")
MEMBER_MULTI_SPLIT = re.compile(r"[,，;；\n]+")

@dataclass
class ImportColumn:
    header: str
    kind: str  # system | custom
    system_key: Optional[str] = None
    field: Optional[dict] = None


@dataclass
class CaseImportError:
    row: int
    message: str


@dataclass
class CaseImportResult:
    created: int
    errors: list[CaseImportError]


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _canonical_header(name: str) -> str:
    key = name.strip()
    return HEADER_ALIASES.get(key, key)


def _trim_trailing_empty(headers: list[str]) -> list[str]:
    out = list(headers)
    while out and not out[-1]:
        out.pop()
    return out


def _normalize_file_headers(raw_row: tuple[Any, ...]) -> list[str]:
    return _trim_trailing_empty([_canonical_header(_cell_str(h)) for h in raw_row])


def _headers_match(file_headers: list[str], expected: list[str]) -> bool:
    return _trim_trailing_empty(file_headers) == expected


def _find_header_row_index(rows: list[tuple[Any, ...]], expected: list[str]) -> Optional[int]:
    for idx, row in enumerate(rows[:20]):
        if _headers_match(_normalize_file_headers(row), expected):
            return idx
    return None


def _parse_priority(raw: str) -> tuple[Optional[CasePriority], Optional[str]]:
    if not raw:
        return CasePriority.P2, None
    upper = raw.upper()
    try:
        return CasePriority(upper), None
    except ValueError:
        return None, f"无效优先级「{raw}」，应为 P0/P1/P2/P3"


def _parse_switch(raw: str) -> tuple[Optional[bool], Optional[str]]:
    if not raw:
        return False, None
    low = raw.strip().lower()
    if low in ("是", "true", "1", "yes", "y"):
        return True, None
    if low in ("否", "false", "0", "no", "n"):
        return False, None
    return None, f"无效的是/否值「{raw}」"


def _parse_date(raw: str, cell_value: Any) -> tuple[Optional[str], Optional[str]]:
    if isinstance(cell_value, datetime):
        return cell_value.date().isoformat(), None
    if isinstance(cell_value, date):
        return cell_value.isoformat(), None
    if not raw:
        return None, None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat(), None
        except ValueError:
            continue
    return None, f"无效日期「{raw}」，请使用 YYYY-MM-DD"


def _sort_template_fields(fields: list) -> list:
    return sorted(fields, key=lambda f: (f.get("sort", 0), f.get("name", "")))


async def _load_fields_schema(db: AsyncSession, project_id: str) -> list:
    tpl = await db.execute(
        select(ProjectFieldTemplate).where(
            ProjectFieldTemplate.project_id == project_id,
            ProjectFieldTemplate.scene == TemplateScene.functional_case,
        )
    )
    row = tpl.scalar_one_or_none()
    if not row:
        return []
    return row.fields or []


def build_import_columns(fields_schema: list) -> list[ImportColumn]:
    system_headers = {h for h, _ in CASE_SYSTEM_IMPORT_COLUMNS}
    cols: list[ImportColumn] = [
        ImportColumn(header=h, kind="system", system_key=k) for h, k in CASE_SYSTEM_IMPORT_COLUMNS
    ]
    for field in _sort_template_fields(fields_schema):
        ftype = field.get("type") or "text"
        if ftype in SKIP_FIELD_TYPES:
            continue
        name = (field.get("name") or field["id"]).strip()
        if name in system_headers or name in CASE_RESERVED_FIELD_NAMES:
            continue
        cols.append(ImportColumn(header=name, kind="custom", field=field))
    return cols


def _build_requirement_lookups(
    rows: list[Requirement],
) -> tuple[dict[str, str], dict[str, str | None]]:
    by_num = {str(r.num): r.id for r in rows}
    by_title: dict[str, str | None] = {}
    for r in rows:
        title = r.title.strip()
        if not title:
            continue
        if title in by_title:
            by_title[title] = None
        else:
            by_title[title] = r.id
    return by_num, by_title


def _resolve_requirement_token(
    token: str,
    *,
    by_num: dict[str, str],
    by_title: dict[str, str | None],
) -> tuple[str | None, Optional[str]]:
    if token.isdigit():
        rid = by_num.get(token)
        if not rid:
            return None, f"未找到需求编号 {token}"
        return rid, None
    rid = by_title.get(token)
    if rid is None and token in by_title:
        return None, f"需求标题「{token}」不唯一，请使用需求编号"
    if rid:
        return rid, None
    return None, f"未找到需求「{token}」，请填写需求标题或编号（如 1）"


async def _parse_requirement_ids_cell(
    db: AsyncSession, project_id: str, raw: str
) -> tuple[list[str], Optional[str]]:
    if not raw.strip():
        return [], None
    result = await db.execute(select(Requirement).where(Requirement.project_id == project_id))
    rows = list(result.scalars().all())
    by_num, by_title = _build_requirement_lookups(rows)
    ids: list[str] = []
    for part in [p.strip() for p in TAG_SPLIT.split(raw) if p.strip()]:
        token = part.lstrip("#").strip()
        rid, err = _resolve_requirement_token(token, by_num=by_num, by_title=by_title)
        if err:
            return [], err
        if rid:
            ids.append(rid)
    return ids, None


def expected_headers(columns: list[ImportColumn]) -> list[str]:
    return [c.header for c in columns]


async def _load_member_lookups(
    db: AsyncSession, project_id: str
) -> tuple[dict[str, str], dict[str, Optional[str]]]:
    result = await db.execute(
        select(User)
        .join(ProjectMember, ProjectMember.user_id == User.id)
        .where(ProjectMember.project_id == project_id)
    )
    users = list(result.scalars().all())
    by_email: dict[str, str] = {}
    by_name: dict[str, Optional[str]] = {}
    for u in users:
        by_email[u.email.lower()] = u.id
        if u.name in by_name:
            by_name[u.name] = None
        else:
            by_name[u.name] = u.id
    return by_email, by_name


def _resolve_member(
    raw: str,
    by_email: dict[str, str],
    by_name: dict[str, Optional[str]],
) -> tuple[Optional[str], Optional[str]]:
    key = raw.strip()
    if not key:
        return None, None
    if "@" in key:
        uid = by_email.get(key.lower())
        if uid:
            return uid, None
        return None, f"未找到成员邮箱「{key}」"
    uid = by_name.get(key)
    if uid:
        return uid, None
    if key in by_name and by_name[key] is None:
        return None, f"成员姓名「{key}」不唯一，请使用邮箱"
    return None, f"未找到成员「{key}」"


def _parse_custom_cell(
    field: dict,
    raw: str,
    cell_value: Any,
    *,
    by_email: dict[str, str],
    by_name: dict[str, Optional[str]],
) -> tuple[Any, Optional[str]]:
    ftype = field.get("type") or "text"
    fname = field.get("name") or field["id"]
    opts = field.get("options") or []

    if ftype in ("text", "textarea"):
        return raw or None, None
    if ftype == "richtext":
        if not raw:
            return None, None
        return {"text": raw, "files": []}, None
    if ftype == "number":
        if not raw and cell_value is None:
            return None, None
        if isinstance(cell_value, (int, float)):
            return cell_value, None
        try:
            return float(raw) if "." in raw else int(raw), None
        except ValueError:
            return None, f"「{fname}」必须是数字"
    if ftype == "date":
        return _parse_date(raw, cell_value)
    if ftype == "switch":
        return _parse_switch(raw)
    if ftype == "select":
        if not raw:
            return None, None
        if opts and raw not in opts:
            return None, f"「{fname}」选项无效，允许：{', '.join(opts)}"
        return raw, None
    if ftype == "multi_select":
        if not raw:
            return [] if not field.get("required") else None, None
        parts = [p.strip() for p in MULTI_SPLIT.split(raw) if p.strip()]
        if opts:
            for p in parts:
                if p not in opts:
                    return None, f"「{fname}」包含无效选项「{p}」"
        return parts, None
    if ftype == "member":
        if not raw:
            return None, None
        return _resolve_member(raw, by_email, by_name)
    if ftype == "member_multi":
        if not raw:
            return [], None
        ids: list[str] = []
        for part in [p.strip() for p in MEMBER_MULTI_SPLIT.split(raw) if p.strip()]:
            uid, err = _resolve_member(part, by_email, by_name)
            if err:
                return None, err
            if uid:
                ids.append(uid)
        return ids, None
    return raw or None, None


def _validation_error_message(exc: HTTPException) -> str:
    detail = exc.detail
    if isinstance(detail, str):
        return detail
    return "字段校验失败"


def _system_example_value(system_key: str) -> str:
    examples = {
        "title": "示例用例标题",
        "priority": "P2",
        "precondition": "前置条件示例",
        "step_text": "步骤示例",
        "expected_result": "预期结果示例",
        "requirement_ids": "示例需求标题",
        "module_path": "父模块-子模块",
    }
    return examples.get(system_key, "")


def _custom_example_value(field: dict) -> str:
    ftype = field.get("type") or "text"
    if ftype == "switch":
        return "否"
    if ftype == "select":
        opts = field.get("options") or []
        return opts[0] if opts else ""
    if ftype == "date":
        return "2026-01-01"
    if ftype == "member":
        return "成员邮箱@example.com"
    if ftype in ("text", "textarea"):
        name = field.get("name") or ""
        return f"{name}示例" if name else "自定义字段示例"
    return ""


def _is_template_example_row(cells: list[str], columns: list[ImportColumn]) -> bool:
    for col, val in zip(columns, cells):
        if col.kind == "system" and col.system_key == "title":
            return val == _system_example_value("title")
    return False


def generate_template_workbook(columns: list[ImportColumn]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "用例导入"
    headers = expected_headers(columns)
    ws.append(headers)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row_is_empty(values: list[str]) -> bool:
    return not any(v for v in values)


def _header_mismatch_detail(file_headers: list[str], expected: list[str]) -> str:
    got = "、".join(file_headers) if file_headers else "（空）"
    exp = "、".join(expected)
    return f"表头与当前项目模板不一致，请重新下载导入模板。期望：{exp}；当前：{got}"


async def parse_import_workbook(
    content: bytes,
    *,
    project_id: str,
    user_id: str,
    db: AsyncSession,
) -> CaseImportResult:
    project_modules = await load_project_modules(db, project_id)
    module_resolver = ImportModuleResolver(
        modules=list(project_modules),
        project_id=project_id,
        db=db,
    )

    fields_schema = await _load_fields_schema(db, project_id)
    columns = build_import_columns(fields_schema)
    expected = expected_headers(columns)

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"无法读取 Excel 文件: {e}",
        ) from e

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel 文件为空")

    header_row_idx = _find_header_row_index(rows, expected)
    if header_row_idx is None:
        first = _normalize_file_headers(rows[0])
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=_header_mismatch_detail(first, expected),
        )

    data_start = header_row_idx + 1
    if data_start < len(rows):
        probe = [_cell_str(c) for c in rows[data_start]] + [""] * len(columns)
        probe = probe[: len(columns)]
        if _is_template_example_row(probe, columns):
            data_start += 1

    member_ids = await load_project_member_user_ids(db, project_id)
    by_email, by_name = await _load_member_lookups(db, project_id)

    created = 0
    errors: list[CaseImportError] = []

    for row_idx, row in enumerate(rows[data_start:], start=data_start + 1):
        cells = [_cell_str(c) for c in row] + [""] * (len(columns) - len(row))
        cells = cells[: len(columns)]
        if _row_is_empty(cells):
            continue

        title = ""
        module_path_raw = ""
        priority = CasePriority.P2
        precondition: Optional[str] = None
        step_text: Optional[str] = None
        expected_result: Optional[str] = None
        requirement_ids: list[str] = []
        custom_fields: dict[str, Any] = {}
        row_error: Optional[str] = None

        for col_index, (col, val) in enumerate(zip(columns, cells)):
            if row_error:
                break
            if col.kind == "system":
                if not col.system_key:
                    continue
                if col.system_key == "module_path":
                    module_path_raw = val
                elif col.system_key == "title":
                    title = val
                    if not title:
                        row_error = "用例标题不能为空"
                elif col.system_key == "priority":
                    p, err = _parse_priority(val)
                    if err:
                        row_error = err
                    elif p:
                        priority = p
                elif col.system_key == "precondition":
                    precondition = val or None
                elif col.system_key == "step_text":
                    step_text = val or None
                elif col.system_key == "expected_result":
                    expected_result = val or None
                elif col.system_key == "requirement_ids":
                    req_ids, err = await _parse_requirement_ids_cell(db, project_id, val)
                    if err:
                        row_error = err
                    else:
                        requirement_ids = req_ids
            elif col.field:
                original = row[col_index] if col_index < len(row) else None
                parsed, err = _parse_custom_cell(
                    col.field,
                    val,
                    original,
                    by_email=by_email,
                    by_name=by_name,
                )
                if err:
                    row_error = err
                elif parsed is not None:
                    custom_fields[col.field["id"]] = parsed

        if row_error:
            errors.append(CaseImportError(row=row_idx, message=row_error))
            continue

        if not title:
            errors.append(CaseImportError(row=row_idx, message="用例标题不能为空"))
            continue

        target_module_id, mod_err = await module_resolver.resolve(module_path_raw)
        if mod_err:
            errors.append(CaseImportError(row=row_idx, message=mod_err))
            continue
        if not target_module_id:
            errors.append(CaseImportError(row=row_idx, message="无法解析模块"))
            continue

        for field in _sort_template_fields(fields_schema):
            ftype = field.get("type") or "text"
            if ftype in SKIP_FIELD_TYPES:
                continue
            name = (field.get("name") or field["id"]).strip()
            if name in {h for h, _ in CASE_SYSTEM_IMPORT_COLUMNS} or name in CASE_RESERVED_FIELD_NAMES:
                continue
            fid = field["id"]
            if field.get("required") and fid not in custom_fields:
                val = custom_fields.get(fid)
                if val is None or val == "" or val == []:
                    errors.append(
                        CaseImportError(row=row_idx, message=f"必填字段「{field.get('name', fid)}」不能为空")
                    )
                    row_error = "required"
                    break
        if row_error:
            continue

        try:
            await validate_requirement_ids(db, project_id, requirement_ids)
        except ValueError as e:
            errors.append(CaseImportError(row=row_idx, message=str(e)))
            continue

        try:
            validated_custom = validate_custom_fields(
                fields_schema,
                custom_fields,
                project_id=project_id,
                project_member_ids=member_ids,
            )
        except HTTPException as exc:
            errors.append(CaseImportError(row=row_idx, message=_validation_error_message(exc)))
            continue

        case = TestCase(
            project_id=project_id,
            module_id=target_module_id,
            title=title[:512],
            priority=priority,
            precondition=precondition,
            step_text=step_text,
            expected_result=expected_result,
            steps=[],
            tags=[],
            custom_fields=validated_custom,
            created_by=user_id,
        )
        db.add(case)
        await db.flush()
        await set_case_requirements(db, case.id, requirement_ids)
        created += 1

    return CaseImportResult(created=created, errors=errors)


async def generate_template_bytes(db: AsyncSession, project_id: str) -> bytes:
    fields_schema = await _load_fields_schema(db, project_id)
    columns = build_import_columns(fields_schema)
    return generate_template_workbook(columns)
