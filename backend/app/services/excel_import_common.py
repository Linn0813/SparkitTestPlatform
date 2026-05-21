from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Optional

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.project import ProjectMember
from app.models.requirement import Requirement
from app.models.template import ProjectFieldTemplate, TemplateScene
from app.models.user import User

SKIP_FIELD_TYPES = frozenset({"attachment", "requirement_link", "plan_link", "version_link", "status", "module"})

TAG_SPLIT = re.compile(r"[,，;；\n]+")
MULTI_SPLIT = re.compile(r"[,，;；\n]+")
MEMBER_MULTI_SPLIT = re.compile(r"[,，;；\n]+")


@dataclass
class ImportColumn:
    header: str
    kind: str  # system | custom
    system_key: Optional[str] = None
    field: Optional[dict] = None


@dataclass
class ImportRowError:
    row: int
    message: str


@dataclass
class ImportResult:
    created: int
    errors: list[ImportRowError]


def cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def canonical_header(name: str, aliases: dict[str, str]) -> str:
    key = name.strip()
    return aliases.get(key, key)


def trim_trailing_empty(headers: list[str]) -> list[str]:
    out = list(headers)
    while out and not out[-1]:
        out.pop()
    return out


def normalize_file_headers(raw_row: tuple[Any, ...], aliases: dict[str, str]) -> list[str]:
    return trim_trailing_empty([canonical_header(cell_str(h), aliases) for h in raw_row])


def headers_match(file_headers: list[str], expected: list[str]) -> bool:
    return trim_trailing_empty(file_headers) == expected


def find_header_row_index(
    rows: list[tuple[Any, ...]], expected: list[str], aliases: dict[str, str]
) -> Optional[int]:
    for idx, row in enumerate(rows[:20]):
        if headers_match(normalize_file_headers(row, aliases), expected):
            return idx
    return None


def expected_headers(columns: list[ImportColumn]) -> list[str]:
    return [c.header for c in columns]


def sort_template_fields(fields: list) -> list:
    return sorted(fields, key=lambda f: (f.get("sort", 0), f.get("name", "")))


async def load_fields_schema(db: AsyncSession, project_id: str, scene: TemplateScene) -> list:
    tpl = await db.execute(
        select(ProjectFieldTemplate).where(
            ProjectFieldTemplate.project_id == project_id,
            ProjectFieldTemplate.scene == scene,
        )
    )
    row = tpl.scalar_one_or_none()
    if not row:
        return []
    return row.fields or []


def build_import_columns(
    system_columns: list[tuple[str, Optional[str]]],
    fields_schema: list,
    *,
    reserved_names: frozenset[str],
) -> list[ImportColumn]:
    system_headers = {h for h, _ in system_columns}
    cols: list[ImportColumn] = [
        ImportColumn(header=h, kind="system", system_key=k) for h, k in system_columns
    ]
    for field in sort_template_fields(fields_schema):
        ftype = field.get("type") or "text"
        if ftype in SKIP_FIELD_TYPES:
            continue
        name = (field.get("name") or field["id"]).strip()
        if name in system_headers or name in reserved_names:
            continue
        cols.append(ImportColumn(header=name, kind="custom", field=field))
    return cols


def build_requirement_lookups(
    rows: list[Requirement],
) -> tuple[dict[str, str], dict[str, str | None]]:
    by_num = {str(r.num): r.id for r in rows}
    by_title: dict[str, str | None] = {}
    for r in rows:
        title = r.title.strip()
        if not title:
            continue
        if title in by_title:
            by_title[title] = None
        else:
            by_title[title] = r.id
    return by_num, by_title


def resolve_requirement_token(
    token: str,
    *,
    by_num: dict[str, str],
    by_title: dict[str, str | None],
) -> tuple[str | None, Optional[str]]:
    if token.isdigit():
        rid = by_num.get(token)
        if not rid:
            return None, f"未找到需求编号 {token}"
        return rid, None
    rid = by_title.get(token)
    if rid is None and token in by_title:
        return None, f"需求标题「{token}」不唯一，请使用需求编号"
    if rid:
        return rid, None
    return None, f"未找到需求「{token}」，请填写需求标题或编号（如 1）"


async def parse_requirement_ids_cell(
    db: AsyncSession, project_id: str, raw: str
) -> tuple[list[str], Optional[str]]:
    if not raw.strip():
        return [], None
    result = await db.execute(select(Requirement).where(Requirement.project_id == project_id))
    rows = list(result.scalars().all())
    by_num, by_title = build_requirement_lookups(rows)
    ids: list[str] = []
    for part in [p.strip() for p in TAG_SPLIT.split(raw) if p.strip()]:
        token = part.lstrip("#").strip()
        rid, err = resolve_requirement_token(token, by_num=by_num, by_title=by_title)
        if err:
            return [], err
        if rid:
            ids.append(rid)
    return ids, None


async def load_member_lookups(
    db: AsyncSession, project_id: str
) -> tuple[dict[str, str], dict[str, Optional[str]]]:
    result = await db.execute(
        select(User)
        .join(ProjectMember, ProjectMember.user_id == User.id)
        .where(ProjectMember.project_id == project_id)
    )
    users = list(result.scalars().all())
    by_email: dict[str, str] = {}
    by_name: dict[str, Optional[str]] = {}
    for u in users:
        by_email[u.email.lower()] = u.id
        if u.name in by_name:
            by_name[u.name] = None
        else:
            by_name[u.name] = u.id
    return by_email, by_name


def resolve_member(
    raw: str,
    by_email: dict[str, str],
    by_name: dict[str, Optional[str]],
) -> tuple[Optional[str], Optional[str]]:
    key = raw.strip()
    if not key:
        return None, None
    if "@" in key:
        uid = by_email.get(key.lower())
        if uid:
            return uid, None
        return None, f"未找到成员邮箱「{key}」"
    uid = by_name.get(key)
    if uid:
        return uid, None
    if key in by_name and by_name[key] is None:
        return None, f"成员姓名「{key}」不唯一，请使用邮箱"
    return None, f"未找到成员「{key}」"


def parse_switch(raw: str) -> tuple[Optional[bool], Optional[str]]:
    if not raw:
        return False, None
    low = raw.strip().lower()
    if low in ("是", "true", "1", "yes", "y"):
        return True, None
    if low in ("否", "false", "0", "no", "n"):
        return False, None
    return None, f"无效的是/否值「{raw}」"


def parse_date(raw: str, cell_value: Any) -> tuple[Optional[str], Optional[str]]:
    if isinstance(cell_value, datetime):
        return cell_value.date().isoformat(), None
    if isinstance(cell_value, date):
        return cell_value.isoformat(), None
    if not raw:
        return None, None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat(), None
        except ValueError:
            continue
    return None, f"无效日期「{raw}」，请使用 YYYY-MM-DD"


def parse_custom_cell(
    field: dict,
    raw: str,
    cell_value: Any,
    *,
    by_email: dict[str, str],
    by_name: dict[str, Optional[str]],
) -> tuple[Any, Optional[str]]:
    ftype = field.get("type") or "text"
    fname = field.get("name") or field["id"]
    opts = field.get("options") or []

    if ftype in ("text", "textarea"):
        return raw or None, None
    if ftype == "richtext":
        if not raw:
            return None, None
        return {"text": raw, "files": []}, None
    if ftype == "number":
        if not raw and cell_value is None:
            return None, None
        if isinstance(cell_value, (int, float)):
            return cell_value, None
        try:
            return float(raw) if "." in raw else int(raw), None
        except ValueError:
            return None, f"「{fname}」必须是数字"
    if ftype == "date":
        return parse_date(raw, cell_value)
    if ftype == "switch":
        return parse_switch(raw)
    if ftype == "select":
        if not raw:
            return None, None
        if opts and raw not in opts:
            return None, f"「{fname}」选项无效，允许：{', '.join(opts)}"
        return raw, None
    if ftype == "multi_select":
        if not raw:
            return [] if not field.get("required") else None, None
        parts = [p.strip() for p in MULTI_SPLIT.split(raw) if p.strip()]
        if opts:
            for p in parts:
                if p not in opts:
                    return None, f"「{fname}」包含无效选项「{p}」"
        return parts, None
    if ftype == "member":
        if not raw:
            return None, None
        return resolve_member(raw, by_email, by_name)
    if ftype == "member_multi":
        if not raw:
            return [], None
        ids: list[str] = []
        for part in [p.strip() for p in MEMBER_MULTI_SPLIT.split(raw) if p.strip()]:
            uid, err = resolve_member(part, by_email, by_name)
            if err:
                return None, err
            if uid:
                ids.append(uid)
        return ids, None
    return raw or None, None


def validation_error_message(exc: HTTPException) -> str:
    detail = exc.detail
    if isinstance(detail, str):
        return detail
    return "字段校验失败"


def generate_template_workbook(columns: list[ImportColumn], *, sheet_title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    headers = expected_headers(columns)
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def row_is_empty(values: list[str]) -> bool:
    return not any(v for v in values)


def header_mismatch_detail(file_headers: list[str], expected: list[str]) -> str:
    got = "、".join(file_headers) if file_headers else "（空）"
    exp = "、".join(expected)
    return f"表头与当前项目模板不一致，请重新下载导入模板。期望：{exp}；当前：{got}"


def is_template_example_row(cells: list[str], columns: list[ImportColumn], example_title: str) -> bool:
    for col, val in zip(columns, cells):
        if col.kind == "system" and col.system_key == "title":
            return val == example_title
    return False


def load_workbook_rows(content: bytes) -> list[tuple[Any, ...]]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=f"无法读取 Excel 文件: {e}"
        ) from e
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel 文件为空")
    return rows


def resolve_header_and_data_start(
    rows: list[tuple[Any, ...]],
    expected: list[str],
    columns: list[ImportColumn],
    aliases: dict[str, str],
    *,
    example_title: str | None = None,
) -> int:
    header_row_idx = find_header_row_index(rows, expected, aliases)
    if header_row_idx is None:
        first = normalize_file_headers(rows[0], aliases)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=header_mismatch_detail(first, expected)
        )
    data_start = header_row_idx + 1
    if example_title and data_start < len(rows):
        probe = [cell_str(c) for c in rows[data_start]] + [""] * len(columns)
        probe = probe[: len(columns)]
        if is_template_example_row(probe, columns, example_title):
            data_start += 1
    return data_start


def check_required_custom_fields(
    fields_schema: list,
    custom_fields: dict[str, Any],
    *,
    system_headers: set[str],
    reserved_names: frozenset[str],
) -> Optional[str]:
    for field in sort_template_fields(fields_schema):
        ftype = field.get("type") or "text"
        if ftype in SKIP_FIELD_TYPES:
            continue
        name = (field.get("name") or field["id"]).strip()
        if name in system_headers or name in reserved_names:
            continue
        fid = field["id"]
        if field.get("required") and fid not in custom_fields:
            val = custom_fields.get(fid)
            if val is None or val == "" or val == []:
                return f"必填字段「{field.get('name', fid)}」不能为空"
    return None
