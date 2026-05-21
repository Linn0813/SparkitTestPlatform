"""Tests for bug import template columns."""

from app.services.bug_import import BUG_HEADER_ALIASES, build_bug_import_columns
from app.services.excel_import_common import expected_headers, generate_template_workbook


def test_build_bug_import_columns():
    schema = [
        {"id": "remark", "name": "备注", "type": "text", "required": False, "options": [], "sort": 0},
        {"id": "dup", "name": "缺陷标题", "type": "text", "required": False, "options": [], "sort": 1},
    ]
    cols = build_bug_import_columns(schema)
    headers = expected_headers(cols)
    assert headers[0] == "缺陷标题"
    assert headers[1] == "状态"
    assert "备注" in headers
    assert "缺陷标题" not in headers[2:]


def test_generate_bug_template_header_only():
    cols = build_bug_import_columns([])
    data = generate_template_workbook(cols, sheet_title="缺陷导入")
    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == "缺陷标题"
    assert "关联需求" in rows[0]
