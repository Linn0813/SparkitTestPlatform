"""Tests for case import template columns and header matching."""

from app.models.requirement import Requirement
from app.services.case_import import (
    CASE_HEADER_ALIASES,
    build_case_import_columns,
    generate_template_bytes,
)
from app.services.excel_import_common import (
    build_requirement_lookups,
    expected_headers,
    find_header_row_index,
    generate_template_workbook,
    headers_match,
    normalize_file_headers,
    resolve_requirement_token,
)


def test_resolve_requirement_by_num_and_title():
    rows = [
        Requirement(project_id="p", num=1, title="Avatar-v1.0", created_by="u"),
        Requirement(project_id="p", num=2, title="登录", created_by="u"),
    ]
    rows[0].id = "r1"
    rows[1].id = "r2"
    by_num, by_title = build_requirement_lookups(rows)

    rid, err = resolve_requirement_token("1", by_num=by_num, by_title=by_title)
    assert err is None and rid == "r1"

    rid, err = resolve_requirement_token("Avatar-v1.0", by_num=by_num, by_title=by_title)
    assert err is None and rid == "r1"

    dup = Requirement(project_id="p", num=3, title="重复", created_by="u")
    dup.id = "r3"
    dup2 = Requirement(project_id="p", num=4, title="重复", created_by="u")
    dup2.id = "r4"
    _, by_dup = build_requirement_lookups([*rows, dup, dup2])
    rid, err = resolve_requirement_token("重复", by_num=by_num, by_title=by_dup)
    assert rid is None and "不唯一" in (err or "")

    rid, err = resolve_requirement_token("不存在", by_num=by_num, by_title=by_title)
    assert rid is None and "未找到需求" in (err or "")


def test_build_import_columns_includes_module_and_custom():
    schema = [
        {"id": "remark", "name": "备注", "type": "textarea", "required": False, "options": [], "sort": 1},
        {"id": "dup", "name": "用例标题", "type": "text", "required": False, "options": [], "sort": 2},
    ]
    cols = build_case_import_columns(schema)
    headers = expected_headers(cols)
    assert headers[0] == "模块"
    assert headers[1] == "用例标题"
    assert "备注" in headers
    assert "用例标题" not in headers[2:]


def test_header_aliases_and_row_detection():
    expected = expected_headers(build_case_import_columns([]))
    alias_row = list(expected)
    alias_row[1] = "标题"
    rows = [tuple(alias_row)]
    idx = find_header_row_index(rows, expected, CASE_HEADER_ALIASES)
    assert idx == 0
    assert headers_match(normalize_file_headers(alias_row, CASE_HEADER_ALIASES), expected)


def test_generate_template_workbook_header_only():
    cols = build_case_import_columns(
        [{"id": "remark", "name": "备注", "type": "text", "required": False, "options": [], "sort": 0}]
    )
    data = generate_template_workbook(cols, sheet_title="用例导入")
    assert len(data) > 100
    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1
    assert list(rows[0])[:3] == ["模块", "用例标题", "优先级"]
    assert "备注" in rows[0]
